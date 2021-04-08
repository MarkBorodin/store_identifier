import csv
import re
import sys
import time
from concurrent.futures import TimeoutError

import pandas as pd
import psycopg2
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from pebble import ProcessPool
from selenium import webdriver
from url_normalize import url_normalize
from usp.tree import sitemap_tree_for_homepage
from usp.web_client.requests_client import RequestsWebClient


class _RequestsWebClient(RequestsWebClient):
    __USER_AGENT = 'Mozilla/5.0'


class DomainsAndSubdomains(object):
    # words_for_shop = [
    #     'shop', 'Shop', 'SHOP', 'store', 'Store', 'STORE', 'pay', 'Pay', 'PAY', 'cart', 'Cart', 'CART',
    #     'buy', 'Buy', 'BUY', 'franken', 'Franken', 'FRANKEN', 'CHF', 'Gutschein', 'Geschenkkarte', 'toys',
    #     'Toys', 'services', 'produ', 'waren' 'goods', 'Gesellschaftsspiele', 'Mädchenbekleidung',
    #     'Männerbekleidung', 'Jungenbekleidung', 'Babybekleidung', 'Kinderschuhe', 'Spielsachen'
    # ]

    words_for_shop = [
        'checkout', 'shopping cart', 'warenkorb', 'korb', 'basket'
    ]

    words_for_goods = [
        'goods', 'produ', 'commodity', 'ware', 'item', 'article', 'artikel', 'objekte', 'object',
        'dienstleistungen', 'services', 'service', 'bedienung', 'CHF', 'buy', 'franken', 'pay'
    ]

    def __init__(self, file, mode='1'): # noqa
        self.file = file
        self.result_file = f'result_{self.file}'
        self.mode = mode
        self.domains = list()
        self.buffer = list()

    def get_domains(self):
        """get url and other data from file"""

        # open and read the file
        df = pd.read_excel(self.file, engine='openpyxl')
        self.domains = df.to_dict(orient='record')

        # create output file
        columns = list(df.columns.values)
        columns.append('shop (Yes/No)')
        columns.append('number of products')

        # # with openpyxl lib
        # wb = Workbook()
        # ws = wb.active
        # ws.append(columns)
        # wb.save(self.result_file)
        # wb.close()

        # with csv lib
        with open(self.result_file, "w", newline="", encoding='UTF-8') as f:
            writer = csv.writer(f)
            writer.writerows([columns])

    @staticmethod
    def clear_url(target):
        """tidy up url"""
        return re.sub('.*www\.', '', target, 1).split('/')[0].strip() # noqa

    @staticmethod
    def normalize_urls_list(common_list):
        """tidy up the url list"""
        lst = []
        for dom in common_list:
            dom = dom.replace('www.', '').replace('http://', '').replace('https://', '')
            lst.append(url_normalize(dom))
        lst = list(set(lst))
        final_list = []
        for link in lst:
            try:
                req = requests.get(link)
                if req.status_code == 200:
                    final_list.append(link)
            except Exception:
                pass
        return lst

    def check_the_quantity_of_goods(self, common_list, item):
        """find the number of products (counting the number of keywords in the links found in the sitemap)"""
        counter = 0
        for link in common_list:
            web_client = _RequestsWebClient()

            tree = sitemap_tree_for_homepage(link, web_client)
            lst = [page.url for page in tree.all_pages()]

            # !!!experimental part!!!
            # 1 way: counting the number of products according to the sitemap links
            if self.mode == '1':
                urls_list = str(lst)
                for word in self.words_for_goods:
                    counter += urls_list.count(word)

            # 2 way: follow each link in sitemap and check keywords on each page
            # (if no goods were found in the way 1)
            # (using requests, since with selenium it will take much longer)
            if self.mode == '2':
                urls_list = str(lst)
                for word in self.words_for_goods:
                    counter += urls_list.count(word)
                if counter == 0 and lst != []:
                    counter = self.check_every_page(lst)

            # 3 way: follow each link in sitemap and check keywords on each page (anyway)
            if self.mode == '3':
                counter = self.check_every_page(lst)
            # !!!experimental part!!!

        return counter

    def check_every_page(self, lst):
        """check each page for product availability (by keywords)"""
        counter = 0
        for url in lst:
            try:
                response = requests.get(url, headers={'User-Agent': 'Mozilla/5.0'})
                text = response.text
                for word in self.words_for_goods:
                    counter += text.count(word)
            except Exception as e:
                print(f'check_every_page: {e}')
        return counter

    def task_done(self, future):
        """this is needed to handle the timeout in multithreaded mode""" # noqa
        try:
            result = future.result()  # noqa (blocks until results are ready)
        except TimeoutError as error:
            self.buffer.append(future.item)
            print("Function took longer than %d seconds" % error.args[1])
        except Exception as error:
            self.buffer.append(future.item)
            print("Function raised %s" % error)

    def start(self):
        """start of the program"""

        # get domains from file
        self.get_domains()

        # create a pool for multi-threaded processing
        with ProcessPool(max_workers=5, max_tasks=10) as pool:
            for i in self.domains:
                future = pool.schedule(self.check_domain, args=[i], timeout=360)
                future.item = i
                future.add_done_callback(self.task_done)

        # add objects to the database with which a connection could not be established
        for item in self.buffer:
            self.write_to_file(item, is_shop=False, number_of_goods=0)
            self.open_db()
            self.cur.execute(
                """INSERT INTO Domains_and_subdomains (
                    DUNS, Handelsregister_Nummer, UID, Internet_Adresse, subdomains, Rechtsform, Filiale_Indikator,
                     Mitarbeiter, Mitarbeiter_Gruppe, is_shop, number_of_goods
                     )
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)""", (
                    item['DUNS'],
                    item['Handelsregister-Nummer'],
                    item['UID'],
                    item['Internet-Adresse'],
                    '',
                    item['Rechtsform'],
                    item['Filiale Indikator'],
                    item['Mitarbeiter'],
                    item['Mitarbeiter Gruppe'],
                    True,
                    0
                )
            )
            self.connection.commit()
            self.close_db()

    def check_domain(self, item):
        """check subdomains and check if url is a store (by keywords)"""
        try:
            domain = item['Internet-Adresse']
            subdomains = []
            subdomains_list = list()
            domain_is_shop = False
            domain = str(domain)
            if domain != 'nan':

                # take a domain
                target = self.clear_url(domain)

                # make a request to an external service
                req = requests.get("https://crt.sh/?q=%.{d}&output=json".format(d=target))

                if req.status_code != 200:
                    print("[X] Information not available!")

                else:

                    for (key, value) in enumerate(req.json()):
                        subdomains.append(value['name_value'])

                    subdomains = sorted(set(subdomains))

                    # select the required subdomains
                    for subdomain in subdomains:
                        if 'shop' in subdomain or 'store' in subdomain:
                            domain_is_shop = True
                            if '\n' in subdomain:
                                s = subdomain.split(sep='\n')
                                for v in s:
                                    if 'shop' in v or 'store' in v:
                                        subdomains_list.append(url_normalize(v))
                                        print(f'subdomain_m: {v}')
                            else:
                                subdomains_list.append(url_normalize(subdomain))
                                print(f'subdomain_o: {subdomain}')

                # check is the domain and subdomains of the store
                if domain_is_shop is False:
                    if len(self.is_shop([domain])) > 0:
                        domain_is_shop = True
                    else:
                        domain_is_shop = False

                if domain_is_shop == True: # noqa
                    # check the quantity of goods
                    common_list = [link for link in subdomains_list]
                    subdomains_list = self.normalize_urls_list(common_list)
                    common_list.append(domain)
                    common_list = self.normalize_urls_list(common_list)
                    counter = self.check_the_quantity_of_goods(common_list, item)
                else:
                    counter = 0

                # first way
                # if counter > 0:
                #     domain_is_shop = True

                self.write_to_file(item, is_shop=domain_is_shop, number_of_goods=counter)
                self.open_db()
                self.cur.execute(
                    """INSERT INTO Domains_and_subdomains (
                    DUNS, Handelsregister_Nummer, UID, Internet_Adresse, subdomains, Rechtsform, Filiale_Indikator,
                     Mitarbeiter, Mitarbeiter_Gruppe, is_shop, number_of_goods
                     )
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)""", (
                        item['DUNS'],
                        item['Handelsregister-Nummer'],
                        item['UID'],
                        item['Internet-Adresse'],
                        str(subdomains_list),
                        item['Rechtsform'],
                        item['Filiale Indikator'],
                        item['Mitarbeiter'],
                        item['Mitarbeiter Gruppe'],
                        domain_is_shop,
                        counter
                    )
                )
                self.connection.commit()
                self.close_db()

            else:
                self.write_to_file(item, is_shop=False, number_of_goods=0)
                self.open_db()
                self.cur.execute(
                    """INSERT INTO Domains_and_subdomains (
                    DUNS, Handelsregister_Nummer, UID, Internet_Adresse, subdomains, Rechtsform, Filiale_Indikator,
                     Mitarbeiter, Mitarbeiter_Gruppe, is_shop, number_of_goods
                     )
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)""", (
                        item['DUNS'],
                        item['Handelsregister-Nummer'],
                        item['UID'],
                        item['Internet-Adresse'],
                        '',
                        item['Rechtsform'],
                        item['Filiale Indikator'],
                        item['Mitarbeiter'],
                        item['Mitarbeiter Gruppe'],
                        False,
                        0
                    )
                )
                self.connection.commit()
                self.close_db()

        except Exception as e:
            print(f'check_domain: {e}')
            self.write_to_file(item, is_shop=False, number_of_goods=0)
            self.open_db()
            self.cur.execute(
                """INSERT INTO Domains_and_subdomains (
                    DUNS, Handelsregister_Nummer, UID, Internet_Adresse, subdomains, Rechtsform, Filiale_Indikator,
                     Mitarbeiter, Mitarbeiter_Gruppe, is_shop, number_of_goods
                     )
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)""", (
                    item['DUNS'],
                    item['Handelsregister-Nummer'],
                    item['UID'],
                    item['Internet-Adresse'],
                    '',
                    item['Rechtsform'],
                    item['Filiale Indikator'],
                    item['Mitarbeiter'],
                    item['Mitarbeiter Gruppe'],
                    False,
                    0
                )
            )
            self.connection.commit()
            self.close_db()

    def is_shop(self, subdomains_list):
        """check if url is a store (by keywords)"""
        try:
            shops = []
            chrome_options = webdriver.ChromeOptions()
            chrome_options.headless = True
            browser = webdriver.Chrome('chromedriver', chrome_options=chrome_options)
            for url in subdomains_list:
                browser.get(url)
                time.sleep(5)

                # first way
                soup = str(BeautifulSoup(browser.page_source, 'lxml'))
                for word in self.words_for_shop:
                    if word in soup:
                        shops.append(url)
                    else:
                        pass

            browser.close()
            print(f'SHOPS FOUND: {shops}')
            return shops
        except Exception as e:
            print(f'is_shop: {e}')
            return []

    def open_db(self):
        """open the database"""
        hostname = '127.0.0.1'
        username = 'parsing_admin'
        password = 'parsing_adminparsing_admin'
        database = 'parsing'
        port = "5444"
        self.connection = psycopg2.connect(  # noqa
            host=hostname,
            user=username,
            password=password,
            dbname=database,
            port=port)
        self.cur = self.connection.cursor()  # noqa

    def close_db(self):
        """close the database"""
        self.cur.close()
        self.connection.close()

    def write_to_file(self, item, is_shop, number_of_goods):
        """write data to file"""
        lst = list(item.values())
        lst.append(is_shop)
        lst.append(number_of_goods)

        # # with openpyxl lib
        # wb = load_workbook(filename=self.result_file)
        # ws = wb.active
        # ws.append(lst)
        # wb.save(self.result_file)
        # wb.close()

        # with csv lib
        with open(self.result_file, "a", newline="", encoding='UTF-8') as f:
            writer = csv.writer(f)
            writer.writerows([lst])


if __name__ == '__main__':
    # get company in command line
    # file = sys.argv[1]
    # mode = sys.argv[2]

    file = 'Batch-Company-Adresses_test.xlsx'
    mode = '2'

    # create object
    obj = DomainsAndSubdomains(file, mode)

    # get data
    obj.start()
